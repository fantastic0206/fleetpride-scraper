from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment
from openpyxl import Workbook
import time
import re
import json
import os

def scrape_product_item_detail_urls_by_brand(part_url, part_text, driver, category_url, brand_url):
    url = f"{part_url},{part_text}"

    driver.get(url)
    time.sleep(3)

    try:
        while True:
            show_btn = driver.find_element(By.TAG_NAME, 'webruntime-app')\
                            .find_element(By.CSS_SELECTOR, 'button[aria-label="Show more values for the Brand facet"]')
            
            driver.execute_script("arguments[0].scrollIntoView();", show_btn)
            show_btn.click()
            time.sleep(2)  # Add a short delay after clicking
    except Exception as e:
        print(f"An error occurred: {e}")
        pass  # Exit the loop when the "show more" button is not found

    parsed_html = BeautifulSoup(driver.page_source, 'html.parser')
    
    try:
        part_type_lists = parsed_html.find('webruntime-app')\
                                     .find('main')\
                                     .find(class_='facets_container')\
                                     .find('c-b2b-fp-quantic-facet')\
                                     .find(class_='slds-has-dividers_around-space')\
                                     .find('fieldset')\
                                     .find('ul')\
                                     .find_all('li')
        
        product_urls = []

        for part_type in part_type_lists:
            brand_text = part_type.text.strip()
            part_count_match = re.search(r'\(([\d,]+)\)', brand_text)
            if part_count_match:
                brand_text = brand_text.replace(part_count_match.group(0), '').replace(",", "%2C").replace('&', '%26')
                # part_count = part_count_match.group(1).replace(',', '')
                # if int(part_count) < 5000:
                product_urls.append(f"https://www.fleetpride.com{category_url}#f-fp_prd_brandname={brand_text}&{brand_url},{part_text}")
        
        return product_urls
    
    except Exception as e:
        print(f"An error occurred: {e}")
        return []

def scrape_product_item_detail_urls_by_part_type(part_url, part_text, driver, category_url):
    url = f"{part_url}#cf-fp_prd_categoryhierarchy={part_text}"
    brand_url = f"cf-fp_prd_categoryhierarchy={part_text}"

    driver.get(url)
    time.sleep(3)

    try:
        while True:
            show_btn = driver.find_element(By.TAG_NAME, 'webruntime-app')\
                            .find_element(By.CSS_SELECTOR, 'button[aria-label="Show more values for the Part Type facet"]')
            
            driver.execute_script("arguments[0].scrollIntoView();", show_btn)
            show_btn.click()
            time.sleep(2)  # Add a short delay after clicking
    except Exception as e:
        print(f"An error occurred: {e}")
        pass  # Exit the loop when the "show more" button is not found

    parsed_html = BeautifulSoup(driver.page_source, 'html.parser')
    
    try:
        part_type_lists = parsed_html.find('webruntime-app')\
                                     .find('main')\
                                     .find(class_='facets_container')\
                                     .find('c-b2b-fp-quantic-category-facet')\
                                     .find(class_='slds-has-dividers_around-space')\
                                     .find('fieldset')\
                                     .find('ul')\
                                     .find_all('li')
        
        product_urls = []

        for part_type in part_type_lists:
            part_text = part_type.text.strip()
            part_count_match = re.search(r'\(([\d,]+)\)', part_text)
            if part_count_match:
                part_text = part_text.replace(part_count_match.group(0), '')  # Remove the count from the text
                part_count = part_count_match.group(1).replace(',', '').replace(",", "%2C").replace('&', '%26')
                if int(part_count) < 5000:
                    product_urls.append(f"{url},{part_text}")
                else:
                    product_urls.extend(scrape_product_item_detail_urls_by_brand(url, part_text, driver, category_url, brand_url))
        
        return product_urls
    
    except Exception as e:
        print(f"An error occurred: {e}")
        return []

def scrape_product_item_detail_urls(category_url, driver, index):
    url = 'https://www.fleetpride.com' + category_url
    print(f"category url => {index + 1} => {url}")

    driver.get(url)
    time.sleep(3)

    try:
        while True:
            show_btn = driver.find_element(By.TAG_NAME, 'webruntime-app')\
                            .find_element(By.CSS_SELECTOR, 'button[aria-label="Show more values for the Part Type facet"]')
            
            driver.execute_script("arguments[0].scrollIntoView();", show_btn)
            show_btn.click()
            time.sleep(2)  # Add a short delay after clicking
    except Exception as e:
        print(f"An error occurred: {e}")
        pass  # Exit the loop when the "show more" button is not found
    
    parsed_html = BeautifulSoup(driver.page_source, 'html.parser')
    
    try:
        part_type_lists = parsed_html.find('webruntime-app')\
                                     .find('main')\
                                     .find(class_='facets_container')\
                                     .find('c-b2b-fp-quantic-category-facet')\
                                     .find(class_='slds-has-dividers_around-space')\
                                     .find('fieldset')\
                                     .find('ul')\
                                     .find_all('li')
        
        product_urls = []

        for part_type in part_type_lists:
            part_text = part_type.text.strip()
            part_count_match = re.search(r'\(([\d,]+)\)', part_text)
            if part_count_match:
                part_text = part_text.replace(part_count_match.group(0), '').replace(",", "%2C").replace('&', '%26')  # Remove the count from the text
                part_count = part_count_match.group(1).replace(',', '')
                if int(part_count) < 5000:
                    product_urls.append(f"{url}#cf-fp_prd_categoryhierarchy={part_text}")
                else:
                    product_urls.extend(scrape_product_item_detail_urls_by_part_type(url, part_text, driver, category_url))

        print(product_urls)

        return product_urls
    
    except Exception as e:
        print(f"An error occurred: {e}")
        return []

def get_product_urls_from_page(driver):
    product_urls = []

    parsed_html = BeautifulSoup(driver.page_source, 'html.parser')
    items = parsed_html.select('webruntime-app c-b2b-fp-result-list c-b2b-fp-result')

    for item in items:
        detail_url = item.select_one('.product-title a').get('href')
        product_urls.append(f"https://www.fleetpride.com{detail_url}")

    return product_urls

def scrape_product_urls(driver, start_urls):
    all_product_urls = []

    for index, start_url in enumerate(start_urls):
        driver.get(start_url)
        time.sleep(3)

        while True:
            all_product_urls.extend(get_product_urls_from_page(driver))

            try:
                next_btn = driver.find_element(By.TAG_NAME, 'webruntime-app').find_element(By.CSS_SELECTOR, 'button[title="Next Page"]')
                if 'button-style-disabled' in next_btn.get_attribute('class'):
                    break
                driver.execute_script("arguments[0].scrollIntoView();", next_btn)
                next_btn.click()
                time.sleep(1)
            except Exception as e:
                print(f"An error occurred: {e}")
                break

    return all_product_urls

def scrape_product_info(driver, url, wb, index):
    print(index, "====>", url)
    driver.get(url)
    time.sleep(1)

    parsed_html = BeautifulSoup(driver.page_source, 'html.parser')

    ws = wb.active

    headers = ['Fleet Pride', 'Product Part Number', 'Product Part Name', 'Image Url']

    # Write headers if the first row is empty
    if ws['A1'].value is None:
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    try:
        fleet_pride_value = parsed_html.find('webruntime-app')\
                                       .find('main')\
                                       .find('c-b2b-fp-pdp-container')\
                                       .find('c-b2b-fp-pdp-product-info')\
                                       .find('div', class_='slds-text-color_weak')\
                                       .text.strip().split('|')[2].replace("MPN #: ", "")
    except AttributeError:
        print("Error: Fleet Pride value not found.")
        return

    try:
        product_image = parsed_html.find('webruntime-app')\
                                   .find('main')\
                                   .find('c-b2b-fp-pdp-container')\
                                   .find('c-b2b-fp-pdp-images')\
                                   .find('c-b2b-fp-carousel-image')\
                                   .find('img')['src']
    except AttributeError:
        print("Error: Product image not found.")
        product_image = "Image not available"

    try:
        product_cross_info_container = parsed_html.find('webruntime-app')\
                                                  .find('main')\
                                                  .find('c-b2b-fp-pdp-product-cross-reference')
    except AttributeError:
        product_cross_info_container = None

    if product_cross_info_container is None:
        print("Product cross table is none")
    else:
        try:
            product_cross_infos = product_cross_info_container\
                                  .find('div', class_='slds-grid')\
                                  .find_all('div', class_='slds-size_6-of-12')
            product_cross_infos = product_cross_infos[2:]
        except AttributeError:
            product_cross_infos = []

        if len(product_cross_infos) == 0:
            print("Product cross information is none")
        else:
            product_cross_info_list = []
            
            for i in range(0, len(product_cross_infos), 2):
                name = product_cross_infos[i].find('span').text.strip()
                number_tags = [tag for tag in product_cross_infos[i + 1].find_all(['a', 'span'])]
                number = ', '.join(tag.text.strip() for tag in number_tags)

                product_cross_info_list.append({"product_name": name, "product_number": number})

            for i, product_cross_info in enumerate(product_cross_info_list):
                product_part_number = product_cross_info['product_number']
                product_part_name = product_cross_info['product_name']

                row = ws.max_row + 1

                # Write other product information to adjacent cells
                ws[f'A{row}'] = fleet_pride_value  # Assigning the extracted value or default value
                ws[f'B{row}'] = product_part_number
                ws[f'C{row}'] = product_part_name
                ws[f'D{row}'] = product_image

                print(f"Product information for {product_part_name} added to the Excel file.")

            start_row = ws.max_row - len(product_cross_info_list) + 1
            end_row = ws.max_row
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)

            # Set the alignment of all cells to center both horizontally and vertically
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

# Path to your ChromeDriver
chrome_driver_path = "./drivers/chromedriver/chromedriver.exe"  # Update this path accordingly

# Optional: specify the path to your Chrome executable if it's not in the default location
chrome_options = Options()
chrome_options.add_argument("--start-maximized")

# Set up the ChromeDriver service
service = Service(chrome_driver_path)

# Initialize the WebDriver
driver = webdriver.Chrome(service=service, options=chrome_options)

# Initialize Excel workbook and sheet
wb = Workbook()
wb.active.title = "Scraped Data"

# Open a webpage
driver.get("https://www.fleetpride.com/shop-all-categories")
time.sleep(3)  # Ensure the page loads completely

# Find the category list elements within the webruntime-app element
parsed_html = BeautifulSoup(driver.page_source, 'html.parser')

category_lists = parsed_html.find('webruntime-app')\
                      .find('main')\
                      .find(class_='slds-wrap')\
                      .find_all(class_='category-wrapper')

file_url = 'product_item_detail_url_by_category_ary.json'

product_item_detail_url_by_category_ary = []

for index, category_list in enumerate(category_lists):
    # Find the anchor tag within the category-wrapper element
    category_url = category_list.find('a')['href']
    
    product_item_detail_urls = scrape_product_item_detail_urls(category_url, driver, index)
    if product_item_detail_urls:
        product_item_detail_url_by_category_ary.extend(product_item_detail_urls)

# Print the results
print(product_item_detail_url_by_category_ary)

# Writing array to JSON file
with open(file_url, 'w') as json_file:
    json.dump(product_item_detail_url_by_category_ary, json_file)
    

print('Your product item detail url by category json file exported successfully!')

# File path to save JSON data
file_path = 'product_item_detail_url.json'

if os.path.isfile(file_path):
    if os.path.getsize(file_path) > 0:
        with open(file_path, 'r') as json_file:
            product_item_detail_url_ary = json.load(json_file)

            for index, product_item_detail_url in enumerate(product_item_detail_url_ary):
                scrape_product_info(driver, product_item_detail_url, wb, index)

else:
    product_item_detail_url_ary = []

    product_item_detail_url_ary = scrape_product_urls(driver, product_item_detail_url_by_category_ary)

    print(product_item_detail_url_ary)

    # Writing array to JSON file
    with open(file_path, 'w') as json_file:
        json.dump(product_item_detail_url_ary, json_file)
        

    print('Your product item detail url json file exported successfully!')

# # Save Excel file
wb.save('product_info.xlsx')
print('Your excel file exported successfully!')

# Close the browser
driver.quit()
